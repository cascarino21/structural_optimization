from openpyxl import Workbook
wb = Workbook()
ws = wb.active

import glob
path = input()
files = glob.glob(f"{path}\\*.pdf")

count = 1
for file in files:
    ws[f"A{count}"].hyperlink = file
    ws[f"A{count}"].value = file
    ws[f"A{count}"].style = "Hyperlink"
    count = count +1

wb.save(f"{path}\\Hyperlinks.xlsx")

